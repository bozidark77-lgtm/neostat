"""
run_pipeline_smoke.py — end-to-end check of the recovered pipeline.

Runs the same path the GUI runs (convert each .xlsx -> .csv, then
analyze.generate_report) against the fixtures from make_fixtures.py, and asserts
the output workbook has the seven expected sheets with the expected irregularity
counts. Exits non-zero on any mismatch so CI fails loudly.

Usage:  python tests/make_fixtures.py && python tests/run_pipeline_smoke.py
"""

import sys
import tempfile
from pathlib import Path

# The report uses Serbian labels (č, š, …). On a console that defaults to a
# non-UTF-8 codepage (Windows cp1252) printing them raises UnicodeEncodeError,
# so force UTF-8 on stdout/stderr where the stream supports it.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import pandas as pd                       # noqa: E402
from openpyxl import load_workbook        # noqa: E402

from convert import convert_xlsx_to_csv   # noqa: E402
from analyze import generate_report       # noqa: E402

FIX = REPO / "tests" / "fixtures"
SUPPLIER_XLSX = FIX / "IZVESTAJ DOBAVLJACA TEST.xlsx"
BREZA_XLSX = FIX / "AMS Team TEST - BREZA.xlsx"

EXPECTED_SHEETS = [
    "REZIME", "NEPRAVILNOSTI", "DUPLIRANI_SATI", "MANJAK_SATI",
    "UPARENO", "DOBAVLJAC_NORMALIZOVANO", "BREZA_INTERVALI",
]

# The fixtures are built to trigger exactly one of each irregularity type.
EXPECTED_ISSUE_COUNTS = {
    "Duplirani sati (isti pogon)": 1,
    "Manjak sati (prijavljeno > BREZA)": 1,
    "Preklapanje na različitim pogonima": 1,
    "Nema u BREZA evidenciji": 1,
    "Pogrešan ID kartice": 1,
}


def fail(msg: str) -> "NoReturn":
    print("SMOKE TEST FAILED: " + msg)
    sys.exit(1)


def main() -> None:
    if not SUPPLIER_XLSX.exists() or not BREZA_XLSX.exists():
        fail("fixtures missing — run `python tests/make_fixtures.py` first.")

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        sup_csv = td / "supplier.csv"
        brz_csv = td / "breza.csv"
        out_xlsx = td / "Nalaz.xlsx"

        # 1) XLSX -> CSV (routes by filename, like the GUI does)
        convert_xlsx_to_csv(str(SUPPLIER_XLSX), str(sup_csv))
        convert_xlsx_to_csv(str(BREZA_XLSX), str(brz_csv))
        for p in (sup_csv, brz_csv):
            if not p.exists() or p.stat().st_size == 0:
                fail("conversion produced an empty CSV: " + p.name)

        # 2) CSV -> styled 7-sheet report (15-min hours tolerance)
        generate_report(str(sup_csv), str(brz_csv), str(out_xlsx))
        if not out_xlsx.exists():
            fail("generate_report did not write " + out_xlsx.name)

        # 3) the report must have exactly the seven expected sheets
        wb = load_workbook(out_xlsx, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if sheets != EXPECTED_SHEETS:
            fail("unexpected sheets.\n  got:      " + str(sheets)
                 + "\n  expected: " + str(EXPECTED_SHEETS))

        # 4) the REZIME (summary) sheet must report the expected irregularity counts
        rezime = pd.read_excel(out_xlsx, sheet_name="REZIME", engine="openpyxl")
        metrics = dict(zip(rezime["metrika"], rezime["vrednost"]))
        for metric, want in EXPECTED_ISSUE_COUNTS.items():
            got = int(metrics.get(metric, -1))
            if got != want:
                fail("REZIME['" + metric + "'] = " + str(got) + ", expected " + str(want))

        # 5) every irregularity must name the plant (HBIS issue C)
        nepr = pd.read_excel(out_xlsx, sheet_name="NEPRAVILNOSTI", engine="openpyxl")
        if "pogon" not in nepr.columns:
            fail("NEPRAVILNOSTI is missing the 'pogon' (plant) column.")
        blank = nepr["pogon"].isna() | (nepr["pogon"].astype(str).str.strip() == "")
        if blank.any():
            fail(str(int(blank.sum())) + " irregularity row(s) have no plant.")

        print("SMOKE TEST OK — 7 sheets, all five irregularity types detected, plant on every finding.")
        print("  sheets: " + ", ".join(sheets))
        for metric in EXPECTED_ISSUE_COUNTS:
            print("  " + metric + ": " + str(int(metrics[metric])))


if __name__ == "__main__":
    main()
