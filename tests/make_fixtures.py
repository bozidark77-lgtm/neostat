"""
make_fixtures.py — generate small, valid sample workbooks for testing.

Writes two .xlsx files into tests/fixtures/ using the EXACT column headers the
app expects, named so convert.py routes them to the right reader:

  * "IZVESTAJ DOBAVLJACA TEST.xlsx"  -> supplier reader. Two per-plant sheets
    ("Izvestaj P1" / "Izvestaj P2") with a "Sati rada" column, so the plant and
    claimed-hours flow through exactly like the real report.
  * "AMS Team TEST - BREZA.xlsx"     -> robust BREZA reader (header below a title row)

The data is synthetic but deliberately exercises every finding analyze.py emits:
a clean match, DUPLIRANI_SATI (duplicated hours, same plant), MANJAK_SATI (a
mid-shift exit so actual < claimed), OVERLAP_DIFFERENT_PLANTS, MISSING_ON_BREZA
and WRONG_CARD_ID. Replace these with your real files for a production run.
"""

from pathlib import Path

from openpyxl import Workbook

FIX = Path(__file__).resolve().parent / "fixtures"
FIX.mkdir(parents=True, exist_ok=True)

# Dates are mm/dd/yyyy, times hh:mm, "Sati rada" = claimed hours — as in the report.
SUP_COLS = [
    "Broj kartice", "Ime i prezime",
    "Datum pocetka rada (format mm/dd/yyyy)", "Vreme pocetka rada (format hh:mm)",
    "Datum zavrsetka rada (format mm/dd/yyyy)", "Vreme zavrsetka rada (format hh:mm)",
    "Sati rada", "Broj NZN-a",
]

# Sheet "Izvestaj P1"  -> plant "P1"
P1_ROWS = [
    # clean: BREZA presence >= claimed -> no finding
    ["12345", "Petar Petrović",  "03/01/2026", "07:00", "03/01/2026", "15:00", 8, "NZN-001"],
    # MANJAK_SATI: claims 8h, BREZA shows a mid-shift exit (6.5h actual)
    ["22222", "Marko Marković",  "03/01/2026", "07:00", "03/01/2026", "15:00", 8, "NZN-002"],
    # DUPLIRANI_SATI: the same 8h row appears twice for the same day/plant
    ["33333", "Jovan Jovanović", "03/01/2026", "07:00", "03/01/2026", "15:00", 8, "NZN-003"],
    ["33333", "Jovan Jovanović", "03/01/2026", "07:00", "03/01/2026", "15:00", 8, "NZN-003"],
    # MISSING_ON_BREZA: no BREZA event for this card/date
    ["99999", "Nikola Nikolić",  "03/01/2026", "08:00", "03/01/2026", "16:00", 8, "NZN-004"],
    # WRONG_CARD_ID: same person/date is in BREZA under a different card
    ["55555", "Ana Anić",        "03/01/2026", "09:00", "03/01/2026", "17:00", 8, "NZN-005"],
    # OVERLAP_DIFFERENT_PLANTS (part 1 of 2): same worker, overlapping, plant P1
    ["77777", "Stefan Stefanović", "03/01/2026", "10:00", "03/01/2026", "14:00", 4, "NZN-006"],
]

# Sheet "Izvestaj P2"  -> plant "P2"
P2_ROWS = [
    # OVERLAP_DIFFERENT_PLANTS (part 2 of 2): same worker/day, overlaps the P1 row, plant P2
    ["77777", "Stefan Stefanović", "03/01/2026", "11:00", "03/01/2026", "15:00", 4, "NZN-007"],
]

# BREZA: "Datum i Vreme" is day-first (dd.mm.yyyy); "Smer" = ULAZ / IZLAZ.
BREZA_COLS = ["Broj kartice", "Ime i prezime", "Kapija", "Smer", "Datum i Vreme"]
BREZA_ROWS = [
    ["12345", "Petar Petrović",  "P1-ULAZ",  "ULAZ",  "01.03.2026 07:00:00"],
    ["12345", "Petar Petrović",  "P1-IZLAZ", "IZLAZ", "01.03.2026 15:10:00"],
    # Marko leaves mid-shift and returns -> two intervals, 3h + 3.5h = 6.5h
    ["22222", "Marko Marković",  "P1-ULAZ",  "ULAZ",  "01.03.2026 07:00:00"],
    ["22222", "Marko Marković",  "P1-IZLAZ", "IZLAZ", "01.03.2026 10:00:00"],
    ["22222", "Marko Marković",  "P1-ULAZ",  "ULAZ",  "01.03.2026 11:30:00"],
    ["22222", "Marko Marković",  "P1-IZLAZ", "IZLAZ", "01.03.2026 15:00:00"],
    ["33333", "Jovan Jovanović", "P1-ULAZ",  "ULAZ",  "01.03.2026 07:00:00"],
    ["33333", "Jovan Jovanović", "P1-IZLAZ", "IZLAZ", "01.03.2026 15:10:00"],
    # Ana is present, but under card 66666 (supplier listed her as 55555)
    ["66666", "Ana Anić",        "P1-ULAZ",  "ULAZ",  "01.03.2026 09:00:00"],
    ["66666", "Ana Anić",        "P1-IZLAZ", "IZLAZ", "01.03.2026 17:00:00"],
    # Stefan present 10:00-15:00 (covers both his overlapping supplier rows)
    ["77777", "Stefan Stefanović", "P1-ULAZ",  "ULAZ",  "01.03.2026 10:00:00"],
    ["77777", "Stefan Stefanović", "P2-IZLAZ", "IZLAZ", "01.03.2026 15:00:00"],
]


def write_supplier(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Izvestaj P1"
    ws1.append(SUP_COLS)
    for r in P1_ROWS:
        ws1.append(r)
    ws2 = wb.create_sheet("Izvestaj P2")
    ws2.append(SUP_COLS)
    for r in P2_ROWS:
        ws2.append(r)
    wb.save(path)


def write_breza(path: Path) -> None:
    """BREZA export shape: a title row, THEN the header row — the robust reader scans for it."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["AMS Team — BREZA izveštaj (test)"])   # title row above the header
    ws.append(BREZA_COLS)
    for r in BREZA_ROWS:
        ws.append(r)
    wb.save(path)


def main() -> None:
    sup = FIX / "IZVESTAJ DOBAVLJACA TEST.xlsx"
    brz = FIX / "AMS Team TEST - BREZA.xlsx"
    write_supplier(sup)
    write_breza(brz)
    print("Wrote fixtures:")
    print("  " + str(sup))
    print("  " + str(brz))


if __name__ == "__main__":
    main()
