"""
analyze.py  —  engine: reconcile IZVEŠTAJ DOBAVLJAČA vs BREZA gate events.

Originally reconstructed from the PyInstaller .pyc of Neostat_Analiza.exe (the
parsing helpers, column aliases, Serbian labels and styling are kept verbatim).

The detection logic was then reworked to fix the errors HBIS reported for May
2026 that the original engine missed (see BUILD_SPEC.md → "Analysis changes"):

  * DUPLIRANI_SATI             — same worker billed for overlapping hours on the
                                 same day at the same plant (e.g. 2× 8h). The
                                 original only checked BREZA and skipped same-plant
                                 overlaps, so these were invisible.
  * MANJAK_SATI                — claimed hours (Sati rada) exceed the worker's
                                 actual BREZA presence for the day (late in, early
                                 out, or an exit/re-entry mid-shift). Replaces the
                                 old per-timestamp LATE_IN / EARLY_OUT checks,
                                 which flagged ~97% of rows as false positives
                                 (rounded supplier times vs exact gate times).
  * pogon (plant) on every finding — taken from the supplier sheet name, which
                                 convert.py now preserves in a 'Pogon' column.

Output: a styled multi-sheet .xlsx (REZIME, NEPRAVILNOSTI, DUPLIRANI_SATI,
MANJAK_SATI, UPARENO, DOBAVLJAC_NORMALIZOVANO, BREZA_INTERVALI).
"""

import argparse
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

DEFAULT_SUPPLIER_CSV = "csv_output/IZVESTAJ DOBAVLJACA 05.2026.csv"
DEFAULT_BREZA_CSV = "csv_output/AMS Team 01.05-25.05.2026 - BREZA.csv"


def _norm_col(s) -> str:
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def _read_csv_auto(path) -> pd.DataFrame:
    # 1. Učitavamo fajl sa python engine-om bez predefinisanog zaglavlja (kao sirove podatke)
    df_raw = pd.read_csv(str(path), sep=None, engine="python", encoding="utf-8-sig", dtype=str, header=None)
    
    if df_raw.empty:
        return df_raw

    # Ključne reči koje pouzdano ukazuju na to da je reč o stvarnom redu sa kolonama
    target_keywords = {"broj kartice", "ime i prezime", "kapija", "smer"}
    header_row_index = 0

    # 2. Skeniramo prvih 15 redova da pronađemo gde počinje stvarna tabela
    for idx, row in df_raw.head(15).iterrows():
        # Normalizujemo sve vrednosti u trenutnom redu radi bezbednog poređenja
        row_norms = [str(val).strip().lower() for val in row.values if pd.notna(val)]
        
        # Ako red sadrži bilo koji od naših ključnih naziva kolona, to je to!
        if any(any(keyword in val for keyword in target_keywords) for val in row_norms):
            header_row_index = idx
            break

    # 3. Ako je pravo zaglavlje pomereno naniže, rekonstruišemo DataFrame
    if header_row_index > 0:
        # Postavljamo pronađeni red kao nazive kolona
        columns_labels = df_raw.iloc[header_row_index].values
        # Uzimamo sve podatke ispod tog reda
        df_cleaned = df_raw.iloc[header_row_index + 1:].copy()
        df_cleaned.columns = columns_labels
        df_cleaned.reset_index(drop=True, inplace=True)
        return df_cleaned
    else:
        # Ako je tabela regularna i počinje od prvog reda (indeks 0)
        df_raw.columns = df_raw.iloc[0].values
        df_raw = df_raw.iloc[1:].reset_index(drop=True)
        return df_raw


def _resolve_columns(df: pd.DataFrame, aliases: dict) -> dict:
    norm_to_real = {_norm_col(c): c for c in df.columns}
    resolved = {}
    for canonical, names in aliases.items():
        hit = None
        for n in names:
            if _norm_col(n) in norm_to_real:
                hit = norm_to_real[_norm_col(n)]
                break
        if hit is None:
            raise ValueError("Nedostaje kolona za '" + canonical + "'. Pronađene kolone: " + str(list(df.columns)))
        resolved[canonical] = hit
    return resolved


def _optional_column(df: pd.DataFrame, *names):
    """Return the real column name for the first alias present, or None."""
    norm_to_real = {_norm_col(c): c for c in df.columns}
    for n in names:
        if _norm_col(n) in norm_to_real:
            return norm_to_real[_norm_col(n)]
    return None


def _parse_datetime_robust(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    out = pd.to_datetime(s, errors="coerce", format="mixed")
    if out.isna().any():
        alt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        out = out.fillna(alt)
    return out


def norm_text(s) -> str:
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def norm_card(v) -> str:
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):   # Excel turns card numbers into floats
        s = s[:-2]
    return s


def parse_plant_from_gate(gate) -> str:
    """First alpha(+digits) token of the gate code, e.g. 'P1-ULAZ' -> 'P1'."""
    s = str(gate).strip().upper()
    m = re.match(r"^([A-ZŠĐŽČĆ]+[0-9]*)", s)
    return m.group(1) if m else s[:20]


def parse_supplier(path) -> pd.DataFrame:
    aliases = {
        "card":     ["Broj kartice", "Broj kartice"],
        "name":     ["Ime I prezime", "Ime i prezime", "Ime i Prezime", "Ime I Prezime"],
        "date_in":  [
            "Datum pocetka rada (format mm/dd/yyyy)", 
            "Datum početka rada (format mm/dd/yyyy)",
            "Datum pocetka rada", "Datum početka rada",
            "Datum pocetka", "Datum početka", "Datum"
        ],
        "time_in":  [
            "Vreme pocetka rada (fomat hh:mm)", "Vreme pocetka rada (format hh:mm)", 
            "Vreme početka rada (fomat hh:mm)", "Vreme početka rada (format hh:mm)",
            "Vreme pocetka rada", "Vreme početka rada", "Vreme pocetka", "Vreme početka"
        ],
        "date_out": [
            "Datum zavrsetka rada (format mm/dd/yyyy)", 
            "Datum završetka rada (format mm/dd/yyyy)",
            "Datum zavrsetka rada", "Datum završetka rada",
            "Datum zavrsetka", "Datum završetka"
        ],
        "time_out": [
            "Vreme zavrsetka rada (fomat hh:mm)", "Vreme zavrsetka rada (format hh:mm)", 
            "Vreme završetka rada (fomat hh:mm)", "Vreme završetka rada (format hh:mm)",
            "Vreme zavrsetka rada", "Vreme završetka rada", "Vreme zavrsetka", "Vreme završetka"
        ],
        "nzn":      ["Broj NZN-a", "Broj NZN", "Broj nzn-a", "Broj nzn", "NZN", "nzn"],
    }
    df = _read_csv_auto(path)
    cols = _resolve_columns(df, aliases)
    
    plant_col = _optional_column(df, "Pogon", "Pogon rada")
    hours_col = _optional_column(df, "Sati rada", "Broj sati", "Sati")

    df = df.dropna(subset=[cols["name"], cols["date_in"], cols["time_in"],
                           cols["date_out"], cols["time_out"]], how="any").copy()

    out = pd.DataFrame()
    out["employee_name"] = df[cols["name"]].astype(str).str.strip()
    out["name_key"] = out["employee_name"].map(norm_text)
    out["card_id"] = df[cols["card"]].map(norm_card)
    out["nzn"] = df[cols["nzn"]].astype(str).str.strip() if cols.get("nzn") else ""
    out["plant"] = df[plant_col].astype(str).str.strip() if plant_col else ""
    out["claimed_hours_raw"] = pd.to_numeric(df[hours_col], errors="coerce") if hours_col else pd.NA
    out["in_time"] = _parse_datetime_robust(df[cols["date_in"]].astype(str).str.strip()
                                            + " " + df[cols["time_in"]].astype(str).str.strip())
    out["out_time"] = _parse_datetime_robust(df[cols["date_out"]].astype(str).str.strip()
                                             + " " + df[cols["time_out"]].astype(str).str.strip())
    out = out.dropna(subset=["in_time", "out_time"])
    
    out.loc[out["out_time"] < out["in_time"], "out_time"] += pd.Timedelta(days=1)
    
    dur_h = (out["out_time"] - out["in_time"]).dt.total_seconds() / 3600.0
    out["claimed_hours"] = pd.to_numeric(out["claimed_hours_raw"], errors="coerce").fillna(dur_h).round(2)
    out = out.drop(columns=["claimed_hours_raw"]).reset_index(drop=True)
    out["row_id"] = range(len(out))
    return out[["row_id", "employee_name", "name_key", "card_id", "nzn", "plant",
                "in_time", "out_time", "claimed_hours"]]


def parse_breza_events(path) -> pd.DataFrame:
    aliases = {
        "card": ["Broj kartice"],
        "name": ["Ime i Prezime", "Ime i prezime"],
        "gate": ["Kapija"],
        "dir":  ["Smer"],
        "dt":   ["Datum i Vreme"],
    }
    df = _read_csv_auto(path)
    cols = _resolve_columns(df, aliases)
    df = df.dropna(subset=[cols["name"], cols["dir"], cols["dt"]], how="any").copy()

    out = pd.DataFrame()
    out["employee_name"] = df[cols["name"]].astype(str).str.strip()
    out["name_key"] = out["employee_name"].map(norm_text)
    out["card_id"] = df[cols["card"]].map(norm_card)
    out["gate"] = df[cols["gate"]].astype(str).str.strip()
    out["plant"] = out["gate"].map(parse_plant_from_gate)

    def _dir(s):
        u = str(s).strip().upper()
        if "ULAZ" in u:
            return "IN"
        if "IZLAZ" in u:
            return "OUT"
        return "OTHER"
    out["direction"] = df[cols["dir"]].map(_dir)
    out["event_time"] = pd.to_datetime(df[cols["dt"]], errors="coerce", dayfirst=True)
    out = out.dropna(subset=["event_time"])
    out = out[out["direction"].isin(["IN", "OUT"])]
    out = out.drop_duplicates(subset=["name_key", "card_id", "direction", "event_time", "gate"])
    out = out.sort_values("event_time").reset_index(drop=True)
    out["event_date"] = out["event_time"].dt.date
    out["event_row_id"] = range(len(out))
    return out[["event_row_id", "employee_name", "name_key", "card_id",
                "gate", "plant", "direction", "event_time", "event_date"]]


def parse_breza(path) -> pd.DataFrame:
    """Pair consecutive IN/OUT events per card into work intervals."""
    ev = parse_breza_events(path)
    rows = []
    for card_id, g in ev.groupby("card_id", sort=True):
        g = g.sort_values("event_time")
        pending_in = None
        for _, e in g.iterrows():
            if e["direction"] == "IN":
                pending_in = e
            elif e["direction"] == "OUT" and pending_in is not None:
                rows.append({
                    "employee_name": pending_in["employee_name"],
                    "name_key": pending_in["name_key"],
                    "card_id": card_id,
                    "in_time": pending_in["event_time"],
                    "out_time": e["event_time"],
                    "in_gate": pending_in["gate"],
                    "out_gate": e["gate"],
                    "plant_in": pending_in["plant"],
                    "plant_out": e["plant"],
                })
                pending_in = None
    cols = ["employee_name", "name_key", "card_id", "in_time", "out_time",
            "in_gate", "out_gate", "plant_in", "plant_out"]
    out = pd.DataFrame(rows, columns=cols)
    if out.empty:
        return out
    out = out.reset_index(drop=True)
    out["duration_h"] = ((out["out_time"] - out["in_time"]).dt.total_seconds() / 3600.0).round(2)
    out["row_id"] = range(len(out))
    return out[["row_id"] + cols + ["duration_h"]]


def detect_supplier_overlaps(sup: pd.DataFrame) -> pd.DataFrame:
    if sup is None or sup.empty:
        return pd.DataFrame()
    rows = []
    work = sup.assign(_day=sup["in_time"].dt.date)
    for (_nk, day), g in work.groupby(["name_key", "_day"], sort=False):
        if len(g) < 2:
            continue
        recs = g.sort_values("in_time").to_dict("records")
        for i in range(len(recs)):
            for j in range(i + 1, len(recs)):
                a, b = recs[i], recs[j]
                if b["in_time"] >= a["out_time"]:
                    continue
                same_plant = norm_text(a["plant"]) == norm_text(b["plant"])
                h1 = float(a["claimed_hours"]) if pd.notna(a["claimed_hours"]) else 0.0
                h2 = float(b["claimed_hours"]) if pd.notna(b["claimed_hours"]) else 0.0
                if same_plant:
                    issue, pogon = "DUPLIRANI_SATI", str(a["plant"])
                    details = ("Duplirani/preklapajući sati na istom pogonu: "
                               + str(round(h1, 2)) + "h + " + str(round(h2, 2)) + "h (ukupno "
                               + str(round(h1 + h2, 2)) + "h za isti dan).")
                else:
                    issue, pogon = "OVERLAP_DIFFERENT_PLANTS", str(a["plant"]) + " / " + str(b["plant"])
                    details = "Preklapanje rada na različitim pogonima za isti dan."
                rows.append({
                    "issue_type": issue,
                    "pogon": pogon,
                    "employee_name": a["employee_name"],
                    "datum": day,
                    "card_supplier": a["card_id"],
                    "in_1": a["in_time"], "out_1": a["out_time"], "hours_1": round(h1, 2),
                    "in_2": b["in_time"], "out_2": b["out_time"], "hours_2": round(h2, 2),
                    "total_hours": round(h1 + h2, 2),
                    "details": details,
                })
    return pd.DataFrame(rows)


def match_supplier_to_breza(sup: pd.DataFrame, brz: pd.DataFrame, tol_min=5):
    tol_h = tol_min / 60.0
    matched = []
    issues = []

    brz = brz.copy()
    if not brz.empty:
        brz["day"] = brz["in_time"].dt.date

    for _, s in sup.iterrows():
        day = s["in_time"].date()
        claimed = float(s["claimed_hours"]) if pd.notna(s["claimed_hours"]) \
            else (s["out_time"] - s["in_time"]).total_seconds() / 3600.0

        same_card = brz[(brz["card_id"] == s["card_id"]) & (brz["day"] == day)] \
            if not brz.empty else brz

        if same_card.empty:
            same_name = brz[(brz["name_key"] == s["name_key"]) & (brz["day"] == day)] \
                if not brz.empty else brz
            if not same_name.empty:
                issues.append({
                    "issue_type": "WRONG_CARD_ID", "pogon": s["plant"],
                    "employee_name": s["employee_name"], "datum": day,
                    "card_supplier": s["card_id"], "card_breza": same_name.iloc[0]["card_id"],
                    "claimed_hours": round(claimed, 2), "actual_hours": None, "hours_diff": None,
                    "n_intervals": int(len(same_name)),
                    "details": "Isti radnik i datum postoje u BREZA, ali sa drugim ID kartice.",
                })
            else:
                issues.append({
                    "issue_type": "MISSING_ON_BREZA", "pogon": s["plant"],
                    "employee_name": s["employee_name"], "datum": day,
                    "card_supplier": s["card_id"], "card_breza": None,
                    "claimed_hours": round(claimed, 2), "actual_hours": None, "hours_diff": None,
                    "n_intervals": 0,
                    "details": "Nema događaja u BREZA za isti ID kartice i isti datum.",
                })
            continue

        actual = (same_card["out_time"] - same_card["in_time"]).dt.total_seconds().sum() / 3600.0
        first_in = same_card["in_time"].min()
        last_out = same_card["out_time"].max()
        n_int = int(len(same_card))
        diff = claimed - actual

        matched.append({
            "pogon": s["plant"], "employee_name": s["employee_name"],
            "card_supplier": s["card_id"], "card_breza": s["card_id"], "datum": day,
            "claimed_in": s["in_time"], "claimed_out": s["out_time"],
            "first_in": first_in, "last_out": last_out,
            "claimed_hours": round(claimed, 2), "actual_hours": round(actual, 2),
            "hours_diff": round(diff, 2), "n_intervals": n_int,
        })

        if diff > tol_h:
            gap_note = (" Radnik je imao " + str(n_int) + " ulaza/izlaza (prekid u smeni).") if n_int > 1 else ""
            issues.append({
                "issue_type": "MANJAK_SATI", "pogon": s["plant"],
                "employee_name": s["employee_name"], "datum": day,
                "card_supplier": s["card_id"], "card_breza": s["card_id"],
                "claimed_hours": round(claimed, 2), "actual_hours": round(actual, 2),
                "hours_diff": round(diff, 2), "n_intervals": n_int,
                "details": ("Prijavljeno " + str(round(claimed, 2)) + "h, a BREZA evidentira "
                            + str(round(actual, 2)) + "h (manjak " + str(round(diff, 2)) + "h)." + gap_note),
            })

    return pd.DataFrame(matched), pd.DataFrame(issues)


def _autosize_worksheet(writer, sheet_name, df, min_width=2, max_width=60):
    ws = writer.sheets.get(sheet_name)
    if ws is None:
        return
    for idx, col in enumerate(df.columns, start=1):
        body = df[col]
        longest = 0 if body.empty else body.map(lambda x: len(str(x)) if pd.notna(x) else 0).max()
        width = max(len(str(col)), int(longest))
        width = min(max(width + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _format_worksheet(writer, sheet_name, df):
    ws = writer.sheets.get(sheet_name)
    if ws is None:
        return
    ws.freeze_panes = "A2"
    last_col = get_column_letter(max(len(df.columns), 1))
    ws.auto_filter.ref = "A1:" + last_col + str(len(df) + 1)
    fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    font = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


_ISSUE_LABELS = {
    "DUPLIRANI_SATI": "DUPLIRANI_SATI",
    "MANJAK_SATI": "MANJAK_SATI",
    "OVERLAP_DIFFERENT_PLANTS": "PREKLAPANJE_RAZLICITI_POGONI",
    "MISSING_ON_BREZA": "NEMA_U_BREZA_EVIDENCIJI",
    "WRONG_CARD_ID": "POGRESAN_ID_KARTICE",
}


def _safe_select(df, cols, rename=None):
    rename = rename or {}
    if df is None or df.empty:
        return pd.DataFrame(columns=[rename.get(c, c) for c in cols])
    return df.reindex(columns=cols).rename(columns=rename)


def generate_report(supplier_csv, breza_csv, out_xlsx, tol_min=5):
    sup = parse_supplier(str(supplier_csv))
    brz = parse_breza(str(breza_csv))

    overlaps = detect_supplier_overlaps(sup)
    matched, match_issues = match_supplier_to_breza(sup, brz, tol_min=tol_min)

    issues = pd.concat([overlaps, match_issues], ignore_index=True, sort=False) \
        if (not overlaps.empty or not match_issues.empty) else pd.DataFrame()

    def _count(code):
        if issues.empty or "issue_type" not in issues:
            return 0
        return int((issues["issue_type"] == code).sum())

    n_matched = 0 if matched.empty else len(matched)
    summary = pd.DataFrame([
        ("Broj redova - izveštaj dobavljača", len(sup)),
        ("Broj intervala - BREZA", len(brz)),
        ("Broj uparenih redova", n_matched),
        ("Ukupno nepravilnosti", 0 if issues.empty else len(issues)),
        ("Duplirani sati (isti pogon)", _count("DUPLIRANI_SATI")),
        ("Manjak sati (prijavljeno > BREZA)", _count("MANJAK_SATI")),
        ("Preklapanje na različitim pogonima", _count("OVERLAP_DIFFERENT_PLANTS")),
        ("Nema u BREZA evidenciji", _count("MISSING_ON_BREZA")),
        ("Pogrešan ID kartice", _count("WRONG_CARD_ID")),
    ], columns=["metrika", "vrednost"])

    if not issues.empty:
        nepr = issues.copy()
        nepr["oznaka_nepravilnosti"] = nepr["issue_type"].map(_ISSUE_LABELS).fillna(nepr["issue_type"])
    else:
        nepr = pd.DataFrame()
    nepravilnosti = _safe_select(
        nepr,
        ["oznaka_nepravilnosti", "pogon", "employee_name", "datum", "details"],
        {"employee_name": "ime_prezime", "datum": "datum", "details": "opis"},
    )

    dups = issues[issues["issue_type"] == "DUPLIRANI_SATI"] if not issues.empty else pd.DataFrame()
    dupl_sheet = _safe_select(
        dups,
        ["pogon", "employee_name", "datum", "in_1", "out_1", "hours_1",
         "in_2", "out_2", "hours_2", "total_hours", "details"],
        {"employee_name": "ime_prezime", "in_1": "ulaz_1", "out_1": "izlaz_1", "hours_1": "sati_1",
         "in_2": "ulaz_2", "out_2": "izlaz_2", "hours_2": "sati_2",
         "total_hours": "ukupno_sati", "details": "opis"},
    )

    short = issues[issues["issue_type"] == "MANJAK_SATI"].copy() if not issues.empty else pd.DataFrame()
    if not short.empty:
        short = short.sort_values("hours_diff", ascending=False)
    manjak_sheet = _safe_select(
        short,
        ["pogon", "employee_name", "datum", "claimed_hours", "actual_hours",
         "hours_diff", "n_intervals", "details"],
        {"employee_name": "ime_prezime", "claimed_hours": "sati_dobavljac",
         "actual_hours": "sati_breza", "hours_diff": "manjak_sati",
         "n_intervals": "broj_prolazaka", "details": "opis"},
    )

    upareno = _safe_select(
        matched,
        ["pogon", "employee_name", "card_supplier", "datum", "claimed_in", "claimed_out",
         "first_in", "last_out", "claimed_hours", "actual_hours", "hours_diff", "n_intervals"],
        {"employee_name": "ime_prezime", "card_supplier": "kartica", "claimed_in": "ulaz_dobavljac",
         "claimed_out": "izlaz_dobavljac", "first_in": "prvi_ulaz_breza", "last_out": "poslednji_izlaz_breza",
         "claimed_hours": "sati_dobavljac", "actual_hours": "sati_breza", "hours_diff": "razlika_sati",
         "n_intervals": "broj_prolazaka"},
    )

    sup_export = _safe_select(
        sup,
        ["row_id", "pogon", "employee_name", "card_id", "nzn", "in_time", "out_time", "claimed_hours"],
        {"row_id": "redni_broj", "employee_name": "ime_prezime", "card_id": "id_kartice",
         "nzn": "broj_nzn", "in_time": "ulaz", "out_time": "izlaz", "claimed_hours": "sati_rada"},
    )
    brz_export = _safe_select(
        brz,
        ["row_id", "employee_name", "card_id", "in_time", "out_time", "duration_h",
         "in_gate", "out_gate"],
        {"row_id": "redni_broj", "employee_name": "ime_prezime", "card_id": "id_kartice",
         "in_time": "ulaz", "out_time": "izlaz", "duration_h": "trajanje_sati",
         "in_gate": "kapija_ulaz", "out_gate": "kapija_izlaz"},
    )

    out_xlsx = Path(out_xlsx)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        for name, frame in [
            ("REZIME", summary),
            ("NEPRAVILNOSTI", nepravilnosti),
            ("DUPLIRANI_SATI", dupl_sheet),
            ("MANJAK_SATI", manjak_sheet),
            ("UPARENO", upareno),
            ("DOBAVLJAC_NORMALIZOVANO", sup_export),
            ("BREZA_INTERVALI", brz_export),
        ]:
            frame.to_excel(writer, sheet_name=name, index=False)
            _autosize_worksheet(writer, name, frame)
            _format_worksheet(writer, name, frame)
    return out_xlsx


def main():
    ap = argparse.ArgumentParser(description="Poređenje BREZA i Izveštaja dobavljača")
    ap.add_argument("--supplier", default=DEFAULT_SUPPLIER_CSV,
                    help="Putanja do CSV fajla izveštaja dobavljača (default: " + DEFAULT_SUPPLIER_CSV + ")")
    ap.add_argument("--breza", default=DEFAULT_BREZA_CSV,
                    help="Putanja do CSV fajla BREZA evidencije (default: " + DEFAULT_BREZA_CSV + ")")
    ap.add_argument("--out", default="Nalaz.xlsx", help="Naziv izlaznog Excel fajla")
    ap.add_argument("--tol", type=int, default=5,
                    help="Tolerancija manjka sati, u minutima (default: 5)")
    args = ap.parse_args()
    out = generate_report(args.supplier, args.breza, args.out, tol_min=args.tol)
    print("Završeno: " + str(Path(out).resolve()))


if __name__ == "__main__":
    main()
